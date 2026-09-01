from openpyxl import load_workbook
from config import OUTPUT_DIR, TEMPLATE_PATH
from transform import curr_metrics, prev_metrics, last_metrics, curr_prev, curr_last, top_medici, top_specialitati, detailed_breakdown
from copy import copy

def generate_dashboard(template_path, current_data, prev_data, last_data, prev_variance, last_year_variance):
    wb = load_workbook(template_path)
    ws = wb['dashboard']
    row_map = {
        "total lei": 2,
        "total euro": 3,
        "total pacienti unici": 4,
        "total pacienti noi": 5,
        "total pacienti platitori": 6,
        "incasare medie": 7,
        "total pacienti neplatitori": 8,
        "total discount": 9,
        "total signal iduna": 10
    }

    for key, row in row_map.items():
        ws[f'D{row}'].value = current_data[key]
        ws[f'E{row}'].value = last_data[key]
        ws[f'F{row}'].value = prev_data[key]
        ws[f'G{row}'].value = last_year_variance[f'diferenta {key}']
        ws[f'H{row}'].value = last_year_variance[f'procent diferenta {key}']
        ws[f'J{row}'].value = prev_variance[f'diferenta {key}']
        ws[f'K{row}'].value = prev_variance[f'procent diferenta {key}']

    fill_detailed_table(ws, detailed_breakdown, 22)
    fill_dynamic_table(ws, top_medici, 17)
    fill_dynamic_table(ws, top_specialitati, 13)

    for r in range(14, 100):
        if ws.cell(row=r, column=2).value == "TOP MEDICI":
            ws.row_dimensions[r].height = 49.8
            break
    
    wb.save(OUTPUT_DIR + "Dashboard_Generated.xlsx")

def fill_dynamic_table(worksheet, data_dict, start_row):
    rows_to_insert = len(data_dict) - 1

    if rows_to_insert > 0:
        worksheet.insert_rows(start_row + 1, rows_to_insert)
        for i in range(start_row + 1, start_row + rows_to_insert + 1):
            for j in range(1, 13):
                source_cell = worksheet.cell(row=start_row, column=j)
                target_cell = worksheet.cell(row=i, column=j)

                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.alignment = copy(source_cell.alignment)

    current_row = start_row
    rank = 1

    for name, metrics in data_dict.items():
        worksheet.row_dimensions[current_row].height = None
        worksheet.cell(row=current_row, column=1).value = rank
        worksheet.cell(row=current_row, column=2).value = name
        worksheet.cell(row=current_row, column=4).value = metrics["current"]
        worksheet.cell(row=current_row, column=5).value = metrics["last"]
        worksheet.cell(row=current_row, column=6).value = metrics["prev"]
        worksheet.cell(row=current_row, column=7).value = metrics["current last abs"]
        worksheet.cell(row=current_row, column=8).value = metrics["current last prt"]
        worksheet.cell(row=current_row, column=10).value = metrics["current previous abs"]
        worksheet.cell(row=current_row, column=11).value = metrics["current previous prt"]

        current_row += 1
        rank += 1
    
        


def inject_triple_block(worksheet, anchor_row, target_row, entity_name, metrics_dict):
    for j in range(2, 13):

        # FIRST ROW (NAME)

        source_cell = worksheet.cell(row=anchor_row, column=j)
        target_cell = worksheet.cell(row=target_row, column=j)

        
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)

    worksheet.cell(row=target_row, column=2).value = entity_name
    worksheet.cell(row=target_row, column=4).value = metrics_dict["current"]
    worksheet.cell(row=target_row, column=5).value = metrics_dict["last"]
    worksheet.cell(row=target_row, column=6).value = metrics_dict["prev"]
    worksheet.cell(row=target_row, column=7).value = metrics_dict["current last abs"]
    worksheet.cell(row=target_row, column=8).value = metrics_dict["current last prt"]
    worksheet.cell(row=target_row, column=10).value = metrics_dict["current previous abs"]
    worksheet.cell(row=target_row, column=11).value = metrics_dict["current previous prt"]

    for j in range(2, 13):

        # SECOND ROW (NR. PACIENTI UNICI)

        source_cell = worksheet.cell(row=anchor_row + 1, column=j)
        target_cell = worksheet.cell(row=target_row + 1, column=j)
        
        
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)

    worksheet.cell(row=target_row + 1, column=2).value = f"{entity_name} NR. PACIENTI UNICI"
    worksheet.cell(row=target_row + 1, column=4).value = metrics_dict["current nr patient"]
    worksheet.cell(row=target_row + 1, column=5).value = metrics_dict["last nr patient"]
    worksheet.cell(row=target_row + 1, column=6).value = metrics_dict["previous nr patient"]
    worksheet.cell(row=target_row + 1, column=7).value = metrics_dict["current last nr pat abs"]
    worksheet.cell(row=target_row + 1, column=8).value = metrics_dict["current last nr pat prt"]
    worksheet.cell(row=target_row + 1, column=10).value = metrics_dict["current previous nr pat abs"]
    worksheet.cell(row=target_row + 1, column=11).value = metrics_dict["current previous nr pat prt"]

    for j in range(2, 13):

        # THIRD ROW (INCASARE/PACIENTI UNICI)

        source_cell = worksheet.cell(row=anchor_row + 2, column=j)
        target_cell = worksheet.cell(row=target_row + 2, column=j)


        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)

    worksheet.cell(row=target_row + 2, column=2).value = f"{entity_name} INCASARE/PACIENTI UNICI"
    worksheet.cell(row=target_row + 2, column=4).value = metrics_dict["current avg patient"]
    worksheet.cell(row=target_row + 2, column=5).value = metrics_dict["last avg patient"]
    worksheet.cell(row=target_row + 2, column=6).value = metrics_dict["previous avg patient"]
    worksheet.cell(row=target_row + 2, column=7).value = metrics_dict["current last avg pat abs"]
    worksheet.cell(row=target_row + 2, column=8).value = metrics_dict["current last avg pat prt"]
    worksheet.cell(row=target_row + 2, column=10).value = metrics_dict["current previous avg pat abs"]
    worksheet.cell(row=target_row + 2, column=11).value = metrics_dict["current previous avg pat prt"]

def fill_detailed_table(worksheet, data_dict, start_row):
    total_rows = 0

    for specialty, data in data_dict.items():
        num_docs = len(data["doctors"])

        if num_docs == 0:
            total_rows += 5

        if num_docs > 0:
            total_rows += 4
            total_rows += (num_docs * 4) + 1

    rows_to_insert = total_rows - 3

    if rows_to_insert > 0:
        worksheet.insert_rows(start_row + 3, rows_to_insert)

    current_row = start_row

    for specialty, data in data_dict.items():
        num_docs = len(data["doctors"])

        inject_triple_block(worksheet, start_row, current_row, specialty, data['metrics'])

        if num_docs == 0:
            current_row += 5

        if num_docs > 0:
            current_row += 4

        for i, (doctor, doc_metrics) in enumerate(list(data["doctors"].items())):
            inject_triple_block(worksheet, start_row, current_row, doctor, doc_metrics)

            if i == num_docs - 1:
                current_row += 5
            else:
                current_row += 4
                

generate_dashboard(TEMPLATE_PATH, curr_metrics, prev_metrics, last_metrics, curr_prev, curr_last)
    